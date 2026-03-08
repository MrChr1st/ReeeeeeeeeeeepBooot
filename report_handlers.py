from datetime import datetime
from tempfile import NamedTemporaryFile

from aiogram import F, Router
from aiogram.filters import Command
from aiogram.types import FSInputFile, KeyboardButton, Message, ReplyKeyboardMarkup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from report_storage import (
    add_event,
    build_opened_rows,
    build_request_rows,
    get_financial_report,
    get_last,
    get_last_24h,
    get_opened_today,
    get_stats,
    get_today,
)

router = Router()


def bottom_menu_kb():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📊 Статистика"), KeyboardButton(text="💵 Фин. отчет")],
            [KeyboardButton(text="👤 Кто открыл"), KeyboardButton(text="📚 Последние")],
            [KeyboardButton(text="📅 Сегодня"), KeyboardButton(text="📄 Excel 24ч")],
        ],
        resize_keyboard=True,
    )


def _style_sheet(ws, title: str, headers: list[str]):
    ws.title = title
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ws.row_dimensions[1].height = 24
    return border


def _autosize(ws):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            val = "" if cell.value is None else str(cell.value)
            widths[cell.column] = max(widths.get(cell.column, 0), min(len(val) + 2, 40))
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _add_rows(ws, rows: list[list], border):
    alt_fill = PatternFill(fill_type="solid", fgColor="F7FBFF")
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if r_idx % 2 == 0:
                cell.fill = alt_fill


def make_excel_report(config):
    events = get_last_24h(config.events_file)
    opened_rows = build_opened_rows(events)
    request_rows = build_request_rows(events)

    wb = Workbook()
    ws1 = wb.active
    border1 = _style_sheet(ws1, "Открытия обмена", ["Время", "Пользователь", "User ID", "Профиль"])
    _add_rows(ws1, [[r["time"], r["user"], r["user_id"], r["profile"]] for r in opened_rows], border1)
    _autosize(ws1)

    ws2 = wb.create_sheet("Заявки")
    border2 = _style_sheet(
        ws2,
        "Заявки",
        ["Время", "Заявка", "Пользователь", "User ID", "Профиль", "Отдаёт", "Получает", "Сумма отдачи", "Сумма получения", "Реквизиты", "Оплата", "Статус", "Оплачено в"],
    )
    _add_rows(
        ws2,
        [[r["time"], r["request_id"], r["user"], r["user_id"], r["profile"], r["from_cur"], r["to_cur"], r["amount_from"], r["amount_to"], r["details"], r["payment"], r["status"], r["paid_at"]] for r in request_rows],
        border2,
    )
    _autosize(ws2)

    ws3 = wb.create_sheet("Сводка")
    border3 = _style_sheet(ws3, "Сводка", ["Показатель", "Значение"])
    fin = get_financial_report(config.events_file)
    summary_rows = [
        ["Период", "Последние 24 часа"],
        ["Открыли обмен", len(opened_rows)],
        ["Новых заявок", fin["new_requests"]],
        ["Оплачено", fin["paid_requests"]],
    ]
    _add_rows(ws3, summary_rows, border3)
    _autosize(ws3)

    tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.close()
    wb.save(tmp.name)
    return tmp.name


@router.message(Command("start"))
async def start_cmd(message: Message):
    await message.answer(
        "📊 Repoorrttt bot online\n\nВыберите действие кнопками снизу.",
        reply_markup=bottom_menu_kb(),
    )


@router.message(Command("help"))
async def help_cmd(message: Message):
    await message.answer(
        "Команды:\n"
        "/last - последние события\n"
        "/today - события за сегодня\n"
        "/stats - статистика\n"
        "/finreport - финансовый отчёт за сегодня\n"
        "/opened - кто и когда открыл обмен сегодня\n"
        "/xlsx24 - Excel отчёт за последние 24 часа",
        reply_markup=bottom_menu_kb(),
    )


@router.message(Command("last"))
@router.message(F.text == "📚 Последние")
async def last_cmd(message: Message, config):
    items = get_last(config.events_file, 10)
    if not items:
        await message.answer("Событий пока нет", reply_markup=bottom_menu_kb())
        return
    text = "📚 Последние события\n\n" + "\n\n".join(f"{item['time']}\n{item['text']}" for item in items)
    await message.answer(text[:4096], reply_markup=bottom_menu_kb())


@router.message(Command("today"))
@router.message(F.text == "📅 Сегодня")
async def today_cmd(message: Message, config):
    items = get_today(config.events_file)
    if not items:
        await message.answer("Сегодня событий пока нет", reply_markup=bottom_menu_kb())
        return
    text = "📅 События за сегодня\n\n" + "\n\n".join(f"{item['time']}\n{item['text']}" for item in items[-20:])
    await message.answer(text[:4096], reply_markup=bottom_menu_kb())


@router.message(Command("stats"))
@router.message(F.text == "📊 Статистика")
async def stats_cmd(message: Message, config):
    stats = get_stats(config.events_file)
    await message.answer(
        "📊 Статистика\n\n"
        f"Всего событий: {stats['total']}\n"
        f"Новых заявок: {stats['new']}\n"
        f"Оплачено: {stats['paid']}\n"
        f"Wallet / срочных: {stats['wallet']}\n"
        f"QR запросы: {stats['qr']}\n"
        f"Открыли обмен: {stats['opened']}",
        reply_markup=bottom_menu_kb(),
    )


@router.message(Command("finreport"))
@router.message(F.text == "💵 Фин. отчет")
async def finreport_cmd(message: Message, config):
    report = get_financial_report(config.events_file)
    from_lines = "\n".join(f"• Отдают: {amount:.8f} {cur}" for cur, amount in sorted(report["from_totals"].items())) or "• Нет данных"
    to_lines = "\n".join(f"• Получают: {amount:.8f} {cur}" for cur, amount in sorted(report["to_totals"].items())) or "• Нет данных"
    text = (
        "💵 Финансовый отчёт за сегодня\n\n"
        f"Новых заявок: {report['new_requests']}\n"
        f"Оплачено: {report['paid_requests']}\n\n"
        "Оборот по валюте отдачи:\n"
        f"{from_lines}\n\n"
        "Оборот по валюте получения:\n"
        f"{to_lines}"
    )
    await message.answer(text[:4096], reply_markup=bottom_menu_kb())


@router.message(Command("opened"))
@router.message(F.text == "👤 Кто открыл")
async def opened_cmd(message: Message, config):
    items = get_opened_today(config.events_file)
    if not items:
        await message.answer("Сегодня никто ещё не открывал обмен", reply_markup=bottom_menu_kb())
        return
    text = "👤 Кто и когда открыл обмен сегодня\n\n" + "\n\n".join(f"{item['time']}\n{item['text']}" for item in items[-20:])
    await message.answer(text[:4096], reply_markup=bottom_menu_kb())


@router.message(Command("xlsx24"))
@router.message(F.text == "📄 Excel 24ч")
async def xlsx24_cmd(message: Message, config):
    file_path = make_excel_report(config)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    document = FSInputFile(file_path, filename=f"report_24h_{timestamp}.xlsx")
    await message.answer_document(document, caption="📊 Excel отчёт за последние 24 часа", reply_markup=bottom_menu_kb())


@router.message(F.chat.type.in_({"group", "supergroup"}))
async def log_group_message(message: Message, config):
    if message.text:
        add_event(config.events_file, message.text)
